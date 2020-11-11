import openpyxl
import configparser
import os,re,datetime
import jieba,imageio
from wordcloud import WordCloud
from pysftp import PySFTP
import shutil

filepath=os.getcwd()+'\\files'
backupPath=filepath+'-backup'

# 读取Excel xlsx
def GetDataList(url,minColumn=0,maxColumn=0):
    data_List=[]
    wb = openpyxl.load_workbook(url)
    sheets = wb.sheetnames
    ws = wb.get_sheet_by_name(sheets[0])
    print("\r最大行数：{0} 最大列数：{1}".format(ws.max_row,ws.max_column))
    maxColumn= maxColumn if maxColumn>0 else ws.max_column
    for row in range(1, ws.max_row+1):
        #print("\r{0}".format(row-1))
        data_List.append(tuple([None if ws.cell(row=row, column=col).value=='' else  ws.cell(row=row, column=col).value for col in range(minColumn+1, maxColumn+1)]))
    
    #做文件移动备份操作
    if not os.path.exists(backupPath):
        os.mkdir(backupPath)
    shutil.move(url,os.path.join(backupPath,os.path.basename(url)))
    
    return data_List  


def load_data(path):
    datas=GetDataList(path,0)
    datas=datas[1:]
    word_dict={}
    word_list=''
    jieba.enable_paddle()

    color_list=['#FF292D','#009099','#2E73E6']#建立颜色数组

    for item in datas:
        try:
            word_str=re.sub("[A-Za-z0-9\[\`\~\!\@\#\$\^\&\*\(\)\=\|\{\}\'\:\;\'\,\，\！\【\】\[\]\.\<\>\/\?\~\。\@\#\\\&\*\%]", "",str(item[0]))
        except:
            word_str=str(item[0])
        
        seg_list=jieba.cut(word_str,cut_all=False)
        #print("Paddle Mode: " + '/'.join(list(seg_list)))

        for word in seg_list:
            if(len(str.strip(word))<2):
                continue
            if (word_dict.get(word)):
                word_dict[word] = word_dict[word] + 1
            else:
                word_dict[word] = 1

    sort_words=sorted(word_dict.items(),key=lambda x:x[1],reverse=True)
    sort_words=sort_words[0:201]
    #print(word_dict)#输出前0-100的词

    new_dic={}
    for index,s in enumerate(sort_words) :
        if(len(s[0])<2):
            continue
        word_list = word_list + ' ' + s[0]
        new_dic[s[0]]=s[1]

    color_mask =imageio.imread("tuoyuan.png")
    # wc = WordCloud(
    #         background_color=None, 
    #         mode="RGBA",  #当参数为“RGBA”并且background_color不为空时，背景为透明。
    #         prefer_horizontal=0.5, # 词语水平方向排版出现的频率，默认 0.9 
    #         #relative_scaling=1,   # 词频和字体大小的关联性
    #         max_words=200,  # 显示最大词数
    #         font_path="simsun.ttc",  # 使用字体
    #         font_step=2,
    #         min_font_size=10,
    #         max_font_size=50,
    #         width=810,
    #         height=510,
    #         collocations=False, #避免重复单词
    #         random_state=42, #随机数
    #         colormap=colormap,#设置颜色
    #         #mask=color_mask  # 图幅宽度
    #         )
    # wc.generate_from_frequencies(new_dic)

    from pyecharts import options as opts
    from pyecharts.charts import Page, WordCloud
    from pyecharts.globals import SymbolType
    from pyecharts.render import make_snapshot
    from snapshot_selenium import snapshot as driver
    worldcloud=(
    WordCloud()
    .add("", sort_words,width='810px', height='510px'
    ,word_size_range=[10, 60]
    ,shape="circle"
    ,rotate_step=90
    ,is_draw_out_of_bound=False
    )
    )
    #,mask_image="tuoyuan.png"
    worldcloud.render("word.html")
    #worldcloud.render(path='xiaoguo.png', pixel_ratio=3)

    from unittest import mock
 
    def get_chrome_driver():
        from selenium import webdriver
        options = webdriver.ChromeOptions()
        options.add_argument("headless")
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-gpu')
        options.add_argument('--disable-dev-shm-usage')
        return webdriver.Chrome(options=options)

    with mock.patch('snapshot_selenium.snapshot.get_chrome_driver', get_chrome_driver):
        # 需要安装 snapshot-selenium 或者 snapshot-phantomjs
        imgurl=os.path.join(imgpath,"wordcloud.png")
        make_snapshot(driver, "word.html", imgurl, pixel_ratio=1)

    #make_snapshot(driver, worldcloud.render(), output_name="xiaoguo.png")

    # import matplotlib.pyplot as plt
    # plt.imshow(myWordCloud)
    # plt.axis("off")#不显示坐标轴
    # plt.show()#显示图片

    # wc.to_file("xiaoguo.png")
    




if __name__ == "__main__":
    cp=configparser.RawConfigParser()
    path = os.path.abspath('seetings.ini')
    if os.path.exists(path):
        cp.read(path,encoding="utf-8-sig")
        imgpath=cp.get("system", "IMG_PATH")
        remote_path=cp.get("sftp", "remote_path")

    # 同步sftp
    sftp=PySFTP()
    filetime=datetime.date.today().strftime("%Y-%m-%d 00:00:00")
    sftp.download('','files',datetime.datetime.strptime(filetime,"%Y-%m-%d %H:%M:%S"))

    print("读取目录：",filepath)
    lists = os.listdir(filepath)
    print("开始扫描目录文件")
    totalfilecount=len(lists)
    for i,f in enumerate(lists):
        load_data(os.path.join(filepath,f));
        print("扫描进度：",(i+1),"/",totalfilecount)
    print("执行完毕")

    


    

